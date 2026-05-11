"""
Exportación de resultados a Excel.

Genera un workbook con:
  - Hoja 'Resumen': conteos por tipo documental.
  - Una hoja por tipo documental con sus entidades.
  - Hoja 'Errores' si hay documentos fallidos.
"""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from schemas import (
    DOC_TYPE_LABELS,
    EXPECTED_ENTITIES,
    DocType,
    DocumentResult,
    humanize_entity_label,
)


# Paleta exacta extraida del logo SinergIA Lab
BRAND_PRIMARY = "0C74C8"   # azul de las figuras del logo
BRAND_ACCENT = "FE6B23"    # naranja del logo
BRAND_LIGHT = "FFF4F4"     # crema del fondo del logo

HEADER_FILL = PatternFill("solid", fgColor=BRAND_PRIMARY)
ALT_FILL = PatternFill("solid", fgColor=BRAND_LIGHT)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
CELL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=BRAND_PRIMARY)
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="666666")

_thin = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def build_excel(results: list[DocumentResult]) -> bytes:
    """Construye el Excel y retorna los bytes listos para descargar."""
    wb = Workbook()

    # Hoja Resumen
    ws = wb.active
    ws.title = "Resumen"
    _build_summary_sheet(ws, results)

    # Hoja por tipo documental
    for doc_type in [DocType.CEDULA, DocType.CAMARA_COMERCIO, DocType.RUT, DocType.POLIZA]:
        subset = [r for r in results if r.doc_type == doc_type and r.error is None]
        if not subset:
            continue
        sheet_name = DOC_TYPE_LABELS[doc_type][:31]  # límite de Excel
        sheet = wb.create_sheet(title=sheet_name)
        _build_doctype_sheet(sheet, subset, doc_type)

    # Hoja Errores
    errors = [r for r in results if r.error]
    if errors:
        sheet = wb.create_sheet(title="Errores")
        _build_errors_sheet(sheet, errors)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_summary_sheet(ws: Worksheet, results: list[DocumentResult]) -> None:
    # Título
    ws["A1"] = "DocuInsight — Reporte de procesamiento"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")

    ws["A2"] = f"SinergIA Lab · Generado {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:C2")

    # Encabezado tabla
    headers = ["Tipo documental", "Cantidad", "% del total"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    # Conteos
    counts: dict[DocType, int] = {}
    for r in results:
        counts[r.doc_type] = counts.get(r.doc_type, 0) + 1

    total = max(len(results), 1)
    row = 5
    for dt in [DocType.CEDULA, DocType.CAMARA_COMERCIO, DocType.RUT, DocType.POLIZA, DocType.DESCONOCIDO]:
        n = counts.get(dt, 0)
        ws.cell(row=row, column=1, value=DOC_TYPE_LABELS[dt]).font = CELL_FONT
        ws.cell(row=row, column=2, value=n).font = CELL_FONT
        ws.cell(row=row, column=3, value=f"{n / total * 100:.1f}%").font = CELL_FONT
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = BORDER
            ws.cell(row=row, column=c).alignment = Alignment(horizontal="center" if c > 1 else "left")
        row += 1

    # Total
    ws.cell(row=row, column=1, value="TOTAL").font = BOLD_FONT
    ws.cell(row=row, column=2, value=total).font = BOLD_FONT
    ws.cell(row=row, column=3, value="100.0%").font = BOLD_FONT
    for c in range(1, 4):
        ws.cell(row=row, column=c).border = BORDER
        ws.cell(row=row, column=c).fill = ALT_FILL
        ws.cell(row=row, column=c).alignment = Alignment(horizontal="center" if c > 1 else "left")

    # Anchos
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14


def _build_doctype_sheet(ws: Worksheet, results: list[DocumentResult], doc_type: DocType) -> None:
    expected = EXPECTED_ENTITIES[doc_type]
    headers = ["Archivo", "Confianza clasificación"] + [humanize_entity_label(e) for e in expected]

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    ws.row_dimensions[1].height = 32

    for row_idx, result in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=result.filename).font = CELL_FONT
        ws.cell(row=row_idx, column=2, value=f"{result.doc_type_confidence:.1%}").font = CELL_FONT

        entity_map = {e.label: e.value for e in result.entities}
        for col_idx, label in enumerate(expected, start=3):
            value = entity_map.get(label, "")
            ws.cell(row=row_idx, column=col_idx, value=value).font = CELL_FONT

        for c in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=c).border = BORDER
            if row_idx % 2 == 0:
                ws.cell(row=row_idx, column=c).fill = ALT_FILL

    # Anchos
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    for i in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 24

    ws.freeze_panes = "C2"


def _build_errors_sheet(ws: Worksheet, results: list[DocumentResult]) -> None:
    headers = ["Archivo", "Error"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    for row_idx, result in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=result.filename).font = CELL_FONT
        ws.cell(row=row_idx, column=2, value=result.error or "").font = CELL_FONT
        for c in range(1, 3):
            ws.cell(row=row_idx, column=c).border = BORDER

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 60
